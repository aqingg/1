"""Tests for catalog workbook → JSON generation."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from rule_create3.catalog import read_rules_catalog
from services.tcd08.rule_generator import read_rules_json, validate_rules


THIS_DIR = Path(__file__).resolve().parent
EXAMPLE_XLSX = THIS_DIR / "tcd08_rule_catalog_example_line39.xlsx"
BASE_JSON = THIS_DIR.parent / "config" / "tcd08_section_rules.json"


class CatalogGenerationTest(unittest.TestCase):
    def test_example_workbook_matches_production_fragments(self) -> None:
        base = read_rules_json(BASE_JSON)
        generated = read_rules_catalog(
            EXAMPLE_XLSX,
            condition_refs=base["condition_refs"],
        )
        validation = validate_rules(generated)
        self.assertTrue(validation.ok, validation.errors)

        text_rule = generated["red_paragraph_text_rules"][0]
        delete_rule = generated["red_paragraph_rules"][0]

        expected_text = next(
            rule
            for rule in base["red_paragraph_text_rules"]
            if rule.get("when_ref") == ["peripheral_has_ufs", "peripheral_no_pas"]
        )
        expected_delete = next(
            rule
            for rule in base["red_paragraph_rules"]
            if rule.get("when_ref") == ["peripheral_has_ufs", "inertial_not_only_one_sma"]
        )

        self.assertEqual(text_rule["section"], expected_text["section"])
        self.assertEqual(text_rule["group"], expected_text["group"])
        self.assertEqual(text_rule["when_ref"], expected_text["when_ref"])
        self.assertEqual(text_rule["replacements"], expected_text["replacements"])

        self.assertEqual(delete_rule["section"], expected_delete["section"])
        self.assertEqual(delete_rule["delete_groups"], expected_delete["delete_groups"])
        self.assertEqual(delete_rule["when_ref"], expected_delete["when_ref"])

        self.assertEqual(
            set(generated["condition_refs"]),
            {"peripheral_has_ufs", "peripheral_no_pas", "inertial_not_only_one_sma"},
        )

    def test_section_delete_sheet_generates_calibration_scope(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        catalog_sheet = workbook.create_sheet("规则清单")
        catalog_sheet.append(["序号", "启用", "规则类型", "章节", "段落名称", "检查字段", "条件", "处理结果"])

        section_sheet = workbook.create_sheet("章节删除")
        section_sheet.append(["启用", "检查字段", "关键词(任一)", "删除章节", "说明"])
        section_sheet.append(["是", "Calibration Scope", "FSR|IDF", "3.2", "Front algorithm"])
        section_sheet.append(["是", "Calibration Scope", "IDF", "3.2.2", "Integrated front"])

        with tempfile.TemporaryDirectory() as temp_dir:
            xlsx_path = Path(temp_dir) / "section_delete.xlsx"
            workbook.save(xlsx_path)

            generated = read_rules_catalog(xlsx_path, condition_refs={})

        calibration_scope = generated.get("calibration_scope", {})
        self.assertEqual(calibration_scope.get("field_label"), "Calibration Scope")
        self.assertEqual(len(calibration_scope.get("delete_when_missing", [])), 2)
        self.assertEqual(
            calibration_scope["delete_when_missing"][0],
            {
                "required_any": ["FSR", "IDF"],
                "sections": ["3.2"],
                "description": "Front algorithm",
            },
        )


if __name__ == "__main__":
    unittest.main()

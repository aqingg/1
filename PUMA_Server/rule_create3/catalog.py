"""Read rule catalog xlsx (规则清单 + 技术映射) and build runtime JSON."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CATALOG_SHEET = "规则清单"
TECH_SHEET = "技术映射"
SECTION_DELETE_SHEET_CANDIDATES = (
    "章节删除",
    "calibration_scope",
    "section_delete",
)
SECTION_RE = re.compile(r"^\d+(?:\.\d+)*$")

CATALOG_HEADERS = [
    "序号",
    "启用",
    "规则类型",
    "章节",
    "段落名称",
    "检查字段",
    "条件",
    "处理结果",
]

OPTIONAL_CATALOG_HEADERS = ("md行",)

TEMPLATE_HEADERS = [
    "模板ID",
    "规则类型",
    "章节",
    "红字段落组",
    "条件引用",
    "删除红段组",
    "清单序号",
]

REPLACEMENT_HEADERS = ["模板ID", "序号", "替换原文", "替换新文本"]

RULE_TYPE_ALIASES = {
    "章节删除": "section_delete",
    "删可选段落": "red_delete",
    "改句子": "text_rewrite",
    "默认删除": "default_delete",
}


class CatalogGenerationError(ValueError):
    """Raised when catalog workbook cannot be converted to runtime JSON."""


def read_workbook_sheets(path: Path) -> dict[str, list[list[Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            name: [
                [cell for cell in row]
                for row in workbook[name].iter_rows(values_only=True)
            ]
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()


def sheet_rows_to_dicts(rows: list[list[Any]], header_row_index: int) -> list[dict[str, str]]:
    if header_row_index >= len(rows):
        return []
    headers = [normalize_cell(value) for value in rows[header_row_index]]
    parsed: list[dict[str, str]] = []
    for raw_row in rows[header_row_index + 1 :]:
        if not any(normalize_cell(value) for value in raw_row):
            continue
        row = {
            headers[index]: normalize_cell(raw_row[index]) if index < len(raw_row) else ""
            for index in range(len(headers))
            if headers[index]
        }
        if any(row.values()):
            parsed.append(row)
    return parsed


def read_rules_catalog(
    path: Path,
    *,
    condition_refs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Load catalog xlsx and return a ``tcd08_section_rules``-shaped dict."""
    sheets = read_workbook_sheets(path)
    catalog_matrix = sheets.get(CATALOG_SHEET, [])
    if not catalog_matrix:
        raise CatalogGenerationError(f"Workbook must contain sheet '{CATALOG_SHEET}': {path}")

    catalog_rows = sheet_rows_to_dicts(catalog_matrix, header_row_index=0)
    used_refs: set[str] = set()
    calibration_scope = parse_section_delete_sheet(sheets)
    templates, replacements = parse_tech_sheet(sheets.get(TECH_SHEET, []))

    if calibration_scope:
        for rule in calibration_scope.get("delete_when_missing", []):
            for ref_name in section_rule_ref_names(rule):
                if ref_name:
                    used_refs.add(ref_name)

    templates_by_catalog_id = {
        int(template["catalog_id"]): template
        for template in templates.values()
        if template.get("catalog_id") not in (None, "")
    }

    rules: dict[str, Any] = {
        "red_paragraph_rules": [],
        "red_paragraph_text_rules": [],
    }

    for row_index, raw_row in enumerate(catalog_rows, start=2):
        row = normalize_catalog_row(raw_row)
        if not row_has_content(row):
            continue
        if not is_enabled(row.get("启用", "")):
            continue

        catalog_id = parse_catalog_id(row.get("序号", ""), row_index=row_index)
        template = templates_by_catalog_id.get(catalog_id)
        if template is None:
            raise CatalogGenerationError(
                f"{CATALOG_SHEET} row {row_index}: no 技术映射 row for 清单序号 {catalog_id}"
            )

        description = build_description(row, template)
        section = str(row.get("章节") or template.get("section") or "").strip()
        if not section:
            raise CatalogGenerationError(f"{CATALOG_SHEET} row {row_index}: missing 章节")

        when_ref = parse_when_ref(template.get("when_ref", ""))
        if when_ref:
            for ref_name in ref_names(when_ref):
                used_refs.add(ref_name)

        rule_type = normalize_rule_type(row.get("规则类型", "") or template.get("rule_type", ""))
        if rule_type == "text_rewrite":
            rules["red_paragraph_text_rules"].append(
                build_text_rewrite_rule(
                    template=template,
                    replacements=replacements.get(template["template_id"], []),
                    section=section,
                    description=description,
                    when_ref=when_ref,
                    row_index=row_index,
                )
            )
            continue

        if rule_type in {"red_delete", "default_delete"}:
            delete_groups = parse_delete_groups(
                template.get("delete_groups", ""),
                label=f"{TECH_SHEET} template {template['template_id']} 删除红段组",
            )
            rule: dict[str, Any] = {
                "description": description,
                "section": section,
                "delete_groups": delete_groups,
            }
            if when_ref:
                rule["when_ref"] = when_ref
            elif rule_type == "default_delete":
                rule["when"] = {}
            else:
                raise CatalogGenerationError(
                    f"{CATALOG_SHEET} row {row_index}: 删可选段落 requires 条件引用 in 技术映射"
                )
            rules["red_paragraph_rules"].append(rule)
            continue

        if rule_type == "section_delete":
            raise CatalogGenerationError(
                f"{CATALOG_SHEET} row {row_index}: 章节删除 is not implemented in catalog v1 yet"
            )

        raise CatalogGenerationError(
            f"{CATALOG_SHEET} row {row_index}: unsupported 规则类型 {row.get('规则类型')!r}"
        )

    if condition_refs is None:
        condition_refs = {}
    missing_refs = sorted(used_refs - set(condition_refs))
    if missing_refs:
        raise CatalogGenerationError(
            "condition_refs missing definitions for: " + ", ".join(missing_refs)
        )

    output: dict[str, Any] = {
        "condition_refs": {name: condition_refs[name] for name in sorted(used_refs)},
    }
    if calibration_scope:
        output["calibration_scope"] = calibration_scope
    if rules["red_paragraph_rules"]:
        output["red_paragraph_rules"] = rules["red_paragraph_rules"]
    if rules["red_paragraph_text_rules"]:
        output["red_paragraph_text_rules"] = rules["red_paragraph_text_rules"]
    return output


def find_first_existing_sheet(
    sheets: dict[str, list[list[Any]]],
    candidates: tuple[str, ...],
) -> list[list[Any]]:
    for name in candidates:
        if name in sheets:
            return sheets[name]
    return []


def split_tokens(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，、;；|]+", value) if part.strip()]


def split_sections(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"[,，、;；|\s]+", value) if part.strip()]


def row_value(row: dict[str, str], *names: str) -> str:
    lowered = {str(key).strip().lower(): str(value).strip() for key, value in row.items()}
    for name in names:
        value = lowered.get(name.strip().lower())
        if value is not None:
            return value
    return ""


def parse_section_delete_sheet(sheets: dict[str, list[list[Any]]]) -> dict[str, Any] | None:
    matrix = find_first_existing_sheet(sheets, SECTION_DELETE_SHEET_CANDIDATES)
    if not matrix:
        return None

    rows = sheet_rows_to_dicts(matrix, header_row_index=0)
    if not rows:
        return None

    field_label = "Calibration Scope"
    delete_when_missing: list[dict[str, Any]] = []

    for row_index, raw_row in enumerate(rows, start=2):
        row = {str(key).strip(): normalize_cell(value) for key, value in raw_row.items()}
        if not any(row.values()):
            continue

        enabled = row_value(row, "启用", "enabled", "enable", "是否启用")
        if enabled and not is_enabled(enabled):
            continue

        row_field_label = row_value(row, "检查字段", "field_label", "field", "字段")
        if row_field_label:
            if field_label != row_field_label and delete_when_missing:
                raise CatalogGenerationError(
                    "章节删除 sheet supports one 检查字段 only; "
                    f"got {field_label!r} and {row_field_label!r}"
                )
            field_label = row_field_label

        required_any_raw = row_value(
            row,
            "required_any",
            "required_scopes",
            "关键词",
            "关键词(任一)",
            "scope关键词",
            "触发关键词",
        )
        when_any_ref_raw = row_value(
            row,
            "when_any_ref",
            "条件引用(任一)",
            "任一条件引用",
        )
        sections_raw = row_value(
            row,
            "sections",
            "section",
            "删除章节",
            "章节",
            "目标章节",
        )
        description = row_value(row, "description", "说明", "规则说明", "处理结果")

        required_any = split_tokens(required_any_raw)
        when_any_ref = parse_when_ref(when_any_ref_raw) if when_any_ref_raw else None
        sections = split_sections(sections_raw)

        if not required_any and not when_any_ref:
            raise CatalogGenerationError(f"章节删除 row {row_index}: required_any/关键词 is required")
        if not sections:
            raise CatalogGenerationError(f"章节删除 row {row_index}: sections/删除章节 is required")

        invalid_sections = [section for section in sections if not SECTION_RE.fullmatch(section)]
        if invalid_sections:
            raise CatalogGenerationError(
                f"章节删除 row {row_index}: invalid section number(s): {invalid_sections}"
            )

        delete_when_missing.append(
            {
                "required_any": required_any,
                "sections": sections,
                "description": description,
                **({"when_any_ref": when_any_ref} if when_any_ref else {}),
            }
        )

    if not delete_when_missing:
        return None

    return {
        "field_label": field_label,
        "delete_when_missing": delete_when_missing,
    }


def parse_tech_sheet(matrix: list[list[Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, str]]]]:
    templates: dict[str, dict[str, Any]] = {}
    replacements: dict[str, list[dict[str, str]]] = {}
    mode = "scan"

    for row_index, raw_row in enumerate(matrix, start=1):
        cells = [normalize_cell(value) for value in raw_row]
        if not any(cells):
            continue

        if cells[0] == "模板ID" and len(cells) > 1 and cells[1] == "规则类型":
            mode = "templates"
            continue
        if cells[0] == "模板ID" and len(cells) > 1 and cells[1] == "序号":
            mode = "replacements"
            continue

        template_id = cells[0] if cells else ""
        if mode == "templates" and template_id:
            templates[template_id] = {
                "template_id": template_id,
                "rule_type": cell_at(cells, 1),
                "section": cell_at(cells, 2),
                "group": cell_at(cells, 3),
                "when_ref": cell_at(cells, 4),
                "delete_groups": cell_at(cells, 5),
                "catalog_id": cell_at(cells, 6),
            }
            continue

        if mode == "replacements" and template_id:
            from_text = cell_at(cells, 2)
            if not from_text:
                raise CatalogGenerationError(
                    f"{TECH_SHEET} row {row_index}: 替换原文 is required for template {template_id}"
                )
            replacements.setdefault(template_id, []).append(
                {
                    "from": from_text,
                    "to": cell_at(cells, 3),
                }
            )

    return templates, replacements


def cell_at(cells: list[str], index: int) -> str:
    return cells[index] if index < len(cells) else ""


def build_text_rewrite_rule(
    *,
    template: dict[str, Any],
    replacements: list[dict[str, str]],
    section: str,
    description: str,
    when_ref: str | list[str] | None,
    row_index: int,
) -> dict[str, Any]:
    if not replacements:
        raise CatalogGenerationError(
            f"{CATALOG_SHEET} row {row_index}: 改句子 requires replacement rows in 技术映射 "
            f"for template {template['template_id']}"
        )
    try:
        group_index = int(str(template.get("group", "")).strip())
    except ValueError as exc:
        raise CatalogGenerationError(
            f"{TECH_SHEET} template {template['template_id']}: 红字段落组 must be an integer"
        ) from exc

    rule: dict[str, Any] = {
        "description": description,
        "section": section,
        "group": group_index,
        "replacements": replacements,
    }
    if when_ref:
        rule["when_ref"] = when_ref
    return rule


def build_description(row: dict[str, str], template: dict[str, Any]) -> str:
    parts = [
        str(row.get("段落名称", "")).strip(),
        str(row.get("条件", "")).strip(),
        str(row.get("处理结果", "")).strip(),
    ]
    text = " | ".join(part for part in parts if part)
    if text:
        return text
    return str(template.get("template_id", "")).strip()


def normalize_catalog_row(raw_row: dict[str, Any]) -> dict[str, str]:
    row: dict[str, str] = {}
    for key, value in raw_row.items():
        header = str(key).strip()
        if header in CATALOG_HEADERS or header in OPTIONAL_CATALOG_HEADERS:
            row[header] = normalize_cell(value)
    return row


def row_has_content(row: dict[str, str]) -> bool:
    return any(value for key, value in row.items() if key != "序号" or value)


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def is_enabled(value: str) -> bool:
    return value.strip().lower() not in {"否", "no", "n", "false", "0", "disabled"}


def parse_catalog_id(value: str, *, row_index: int) -> int:
    text = normalize_cell(value)
    if not text:
        raise CatalogGenerationError(f"{CATALOG_SHEET} row {row_index}: 序号 is required")
    try:
        return int(text)
    except ValueError as exc:
        raise CatalogGenerationError(
            f"{CATALOG_SHEET} row {row_index}: 序号 must be an integer, got {value!r}"
        ) from exc


def normalize_rule_type(value: str) -> str:
    normalized = value.strip()
    if normalized in RULE_TYPE_ALIASES:
        return RULE_TYPE_ALIASES[normalized]
    lowered = normalized.lower()
    for label, canonical in RULE_TYPE_ALIASES.items():
        if label.lower() == lowered:
            return canonical
    return normalized


def parse_when_ref(value: str) -> str | list[str] | None:
    text = value.strip()
    if not text:
        return None
    parts = [part.strip() for part in re.split(r"[+|,，/\\]+", text) if part.strip()]
    if not parts:
        return None
    return parts[0] if len(parts) == 1 else parts


def ref_names(when_ref: str | list[str]) -> list[str]:
    if isinstance(when_ref, str):
        return [when_ref] if when_ref else []
    return [str(item).strip() for item in when_ref if str(item).strip()]


def section_rule_ref_names(rule: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for key in ("when_ref", "when_any_ref"):
        value = rule.get(key)
        if isinstance(value, str):
            names.extend(ref_names(value))
        elif isinstance(value, list):
            names.extend(ref_names(value))
    return names


def parse_delete_groups(value: str, *, label: str) -> list[int]:
    text = value.strip()
    if not text:
        raise CatalogGenerationError(f"{label} is required for 删可选段落")
    groups: list[int] = []
    for part in re.split(r"[,，、\s]+", text):
        part = part.strip()
        if not part:
            continue
        try:
            groups.append(int(part))
        except ValueError as exc:
            raise CatalogGenerationError(f"{label} must contain integers only: {value!r}") from exc
    if not groups:
        raise CatalogGenerationError(f"{label} is required for 删可选段落")
    return groups
